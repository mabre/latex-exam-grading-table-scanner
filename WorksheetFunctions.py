import re
import tempfile
from typing import List

import cv2
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def write_image_to_cell(ws: Worksheet, image_array: np.array, row: int, col: int):
    with tempfile.NamedTemporaryFile(prefix="scantool_", suffix=".png", delete=False) as tmp_file:
        cv2.imwrite(tmp_file.name, image_array)
        img = Image(tmp_file.name)

    old_height = img.height
    img.height = 20
    img.width = int(img.width * (20 / old_height))

    cell_address = get_column_letter(col) + str(row)

    ws.add_image(img, cell_address)


def column_index_by_title(ws: Worksheet, column_name: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            return col
    raise ValueError(f"Column {column_name} not found in worksheet")


def column_indices_by_title(ws: Worksheet, column_name_regex: str) -> List[int]:
    pattern = re.compile(column_name_regex)
    return [col for col in range(1, ws.max_column + 1) if pattern.match(ws.cell(row=1, column=col).value)]


def column_letter_by_title(ws: Worksheet, column_name: str) -> str:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            return get_column_letter(col)
    raise ValueError(f"Column {column_name} not found in worksheet")
