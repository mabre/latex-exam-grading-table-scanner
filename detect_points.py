import concurrent
import concurrent.futures
import logging
import os
import shutil
import sys
import time
from pathlib import Path
from typing import Optional, Dict, Tuple, Callable, List, Iterable, Any, Union

import cv2
import cv2.aruco as aruco
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from playsound import playsound
from tqdm import tqdm

from GradingTable import GradingTable
from constants import EXERCISE_HEADER_PREFIX, STUDENT_ID_HEADER, SUM_RECOGNIZED_HEADER, SUM_WORKSHEET_HEADER, \
    MAX_CAMERA_IMAGE_PREVIEW_SIZE
from log_setup import logger

ARUCO_TOP_LEFT_ID = 0
ARUCO_BOTTOM_LEFT_ID = 1
ARUCO_BOTTOM_RIGHT_ID = 2
ARUCO_TOP_RIGHT_ID = 3
ARUCO_IDS = [ARUCO_TOP_LEFT_ID, ARUCO_BOTTOM_LEFT_ID, ARUCO_BOTTOM_RIGHT_ID, ARUCO_TOP_RIGHT_ID]
NUM_ARUCO_MARKERS = len(ARUCO_IDS)

ARUCO_CORNER_TOP_LEFT = 0
ARUCO_CORNER_TOP_RIGHT = 1
ARUCO_CORNER_BOTTOM_RIGHT = 2
ARUCO_CORNER_BOTTOM_LEFT = 3

COLORS = {
    -1: (0, 0, 255),  # Red
    0: (0, 0, 255),  # Red
    1: (0, 0, 255),  # Red
    2: (0, 165, 255),  # Orange
    3: (0, 255, 255),  # Yellow
    4: (47, 255, 173),  # Light Green
    5: (0, 255, 0)  # Green
}

def log_execution_time(func: Callable):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        logger.info(f"Execution time for {func.__name__}: {elapsed_time:.4f} seconds")
        return result
    return wrapper


def find_grading_table_and_student_number(frame_data: Tuple[int, np.array]) -> Optional[Tuple[str, np.array, int, int]]:
    """
    :param frame_data: frame number and frame contents
    :return: student number, frame contents, frame number, and number of aruco markers, if all but one aruco markers and a qr code are found; None otherwise
    """
    frame_number, frame = frame_data
    # print(frame_number, number_of_aruco_markers(frame))
    markers_found = number_of_aruco_markers(frame)
    if markers_found >= NUM_ARUCO_MARKERS - 1:
        student_number, _ = student_number_from_qr_code(frame)
        logger.debug(f"Found {markers_found} aruco markers in frame {frame_number}")
        if student_number is not None:
            logger.debug(f"Found student number {student_number} and all aruco markers in frame {frame_number}")
            return student_number, frame, frame_number, markers_found
    return None


def extract_frames_interactively(video_path: str):
    """
    :param video_path: may also be an integer (string) identifying the first, second, ... camera
    """
    relevant_frames = {}
    frame_number = 0
    last_good_frame = -1000
    last_best_color_index = 0
    previous_student_number = None
    new_frames = []
    if video_path.isnumeric():
        cap = cv2.VideoCapture(int(video_path))
    else:
        cap = cv2.VideoCapture(video_path)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1) # limit the buffer so that processing a frame too slowly does not cause the video to lag behind; this has no influence if a video file is used

    if not cap.isOpened():
        logger.error("Could not open video capture")
        return

    try:
        logger.info("Press 'q' to quit capturing")
        while True:
            ret, frame = cap.read()

            if not ret:
                logger.error("Could not read frame, aborting reading new frames (this is normal at the end of a video file)")
                break

            resized_frame = resize(frame, 2000)  # aruco and qr detection seems to have problems with very big resolutions

            aruco_corners, aruco_ids = detect_aruco_markers(resized_frame)
            student_number, qr_corners = student_number_from_qr_code(resized_frame)

            color_index = len(aruco_ids) + 1 if qr_corners is not None else -1 # so we get any green only with detected qr code
            color = COLORS[color_index]

            img_with_arucos = aruco.drawDetectedMarkers(resized_frame.copy(), aruco_corners, aruco_ids, borderColor=color)
            img_with_arucos_qr = draw_qr_code_bounding_box(img_with_arucos, qr_corners, color)

            if len(aruco_ids) >= NUM_ARUCO_MARKERS - 1 and student_number is not None:
                if previous_student_number != student_number:
                    playsound("sound/bell.ogg", block=False)
                    if previous_student_number is not None:
                        relevant_frames[previous_student_number] = get_best_frame(new_frames)
                        new_frames = []
                        last_best_color_index = 0
                previous_student_number = student_number
                new_frames.append((resized_frame, len(aruco_ids)))
                last_good_frame = frame_number
                last_best_color_index = max(color_index, last_best_color_index)

            if frame_number - last_good_frame < 10:
                cv2.rectangle(img_with_arucos_qr, (0, 0), (img_with_arucos_qr.shape[1], 50), COLORS[last_best_color_index], -1)

            frame_number += 1

            cv2.imshow("Detected Markers", resize(img_with_arucos_qr, MAX_CAMERA_IMAGE_PREVIEW_SIZE))

            # quit using 'q'
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        if len(new_frames) > 0 and previous_student_number is not None and previous_student_number not in relevant_frames:
            relevant_frames[previous_student_number] = get_best_frame(new_frames)

        cap.release()
        cv2.destroyAllWindows()
    except Exception as e:
        logger.error(e)
        logger.error("An unexpected error occurred, trying to process the frames already recorded")

    return relevant_frames


@log_execution_time
def extract_frames(video_path: str) -> Dict[int, np.array]:
    """
    reads each frame in the video and considers all frames where a student number (numeric) qr code is found and at least three aruco markers are detected

    For each student number, exactly one frame (the one in the middle) is returned, preferring those where four aruco markers are found.
    """
    # todo safe all frames for each student, than take best sum-matching prediction; Achtung: immer noch die äußeren Frames verwerfen, falls die zu ner anderen Klausur gehören und wir mind. 5 Frames haben
    cap = cv2.VideoCapture(video_path)
    relevant_frames = {}
    frame_number = 0
    previous_student_number = None
    new_frames = []
    futures = []
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    with concurrent.futures.ThreadPoolExecutor() as executor:
        with tqdm(total=frame_count, desc="Loading frames") as pbar:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break

                resized_frame = resize(frame, 2000) # aruco and qr detection seems to have problems with very big resolutions

                # find_grading_table_and_student_number((frame_number, resized_frame))
                future = executor.submit(find_grading_table_and_student_number, (frame_number, resized_frame))
                futures.append(future)
                frame_number += 1
                pbar.update(1)

        for future in tqdm(futures, desc="Detecting grading tables and qr codes"):
            result = future.result()
            if result:
                student_number, frame, frame_number, number_of_arucos = result
                # for debugging purposes: save all frames
                # relevant_frames[student_number * 100_000 + frame_number] = frame
                if previous_student_number != student_number and previous_student_number is not None:
                    relevant_frames[previous_student_number] = get_best_frame(new_frames)
                    new_frames = []
                previous_student_number = student_number
                new_frames.append((frame, number_of_arucos))

    if len(new_frames) > 0 and previous_student_number is not None and previous_student_number not in relevant_frames:
        relevant_frames[previous_student_number] = get_best_frame(new_frames)

    cap.release()

    assert frame_number != 0, f"No frames found in video, is the path {video_path} correct?"

    return relevant_frames


def get_best_frame(frames_with_aruco_count: list[Tuple[np.array, int]]) -> np.array:
    """The best frame is usually the one in the middle of the video, but we prefer those with all aruco markers"""
    # split the list in those where the second tuple element is 4 and those where it is not
    all_aruco_frames = [frame_with_count[0] for frame_with_count in frames_with_aruco_count if frame_with_count[1] == NUM_ARUCO_MARKERS]
    if len(all_aruco_frames) > 0:
        return all_aruco_frames[len(all_aruco_frames) // 2]
    # otherwise, there are only frames with three detected markers
    return frames_with_aruco_count[len(frames_with_aruco_count) // 2][0]


def student_number_from_qr_code(image: np.array) -> Tuple[Optional[str], Optional[np.array]]:
    data, points = read_qr_code(image)

    if points is not None:
        logger.debug(f"QR Code Data: {data}")
        if data.strip() == "":
            logger.debug("Empty QR code data")
            return None, None
        return data, points
    else:
        logger.debug("No QR codes found")
        return None, None


def read_qr_code(image: np.array) -> Tuple[Any, Optional[np.array]]:
    qr_decoder = cv2.QRCodeDetector()
    data, points, _ = qr_decoder.detectAndDecode(image)
    return data, points


def number_of_aruco_markers(image: np.array) -> int:
    corners, _ = detect_aruco_markers(image)
    return len(corners)


def resize(image: np.array, max_length: int) -> np.array:
    h, w = image.shape[:2]
    if max(h, w) > max_length:
        scale_divisor = max(h, w) // max_length + 1
        return cv2.resize(image, (w // scale_divisor, h // scale_divisor))
    return image


def detect_aruco_markers(image: np.array) -> Tuple[Tuple, np.array]:
    # note: use https://chev.me/arucogen/ to generate markers
    aruco_dict = aruco.getPredefinedDictionary(aruco.DICT_4X4_50)
    parameters = aruco.DetectorParameters()
    parameters.cornerRefinementMethod = aruco.CORNER_REFINE_SUBPIX
    aruco_detector = aruco.ArucoDetector(aruco_dict, parameters)

    corners, ids, _rejected_candidates = aruco_detector.detectMarkers(image)
    # debug_draw_aruco_markers(corners, ids, image)

    if ids is None:
        return corners, np.asarray([])

    valid_indices = [i for i, marker_id in enumerate(ids) if marker_id in ARUCO_IDS]
    valid_corners = tuple(corners[i] for i in valid_indices)
    valid_ids = np.array([ids[i] for i in valid_indices])

    if len(np.unique(valid_ids.flatten())) != len(valid_ids):
        logger.debug(f"Found duplicated aruco {valid_ids}, discarding this frame")
        return (), np.asarray([])

    return valid_corners, valid_ids


def de_skew_and_crop_image(image: np.array) -> Optional[np.array]:
    """
    Cuts out the grading table from the image bases on the position of at least three detected aruco markers.

    The returned image is in RGB, and has a left and right margin which is as wide as a digit cell.
    """
    corners, ids = detect_aruco_markers(image)

    if len(ids) >= NUM_ARUCO_MARKERS - 1:
        top_left_index = np.where(ids == ARUCO_TOP_LEFT_ID)[0][0] if ARUCO_TOP_LEFT_ID in ids else None
        bottom_left_index = np.where(ids == ARUCO_BOTTOM_LEFT_ID)[0][0] if ARUCO_BOTTOM_LEFT_ID in ids else None
        bottom_right_index = np.where(ids == ARUCO_BOTTOM_RIGHT_ID)[0][0] if ARUCO_BOTTOM_RIGHT_ID in ids else None
        top_right_index = np.where(ids == ARUCO_TOP_RIGHT_ID)[0][0] if ARUCO_TOP_RIGHT_ID in ids else None

        detected_corners = [None, None, None, None]

        if top_left_index is not None:
            detected_corners[ARUCO_TOP_LEFT_ID] = corners[top_left_index][0][ARUCO_CORNER_BOTTOM_RIGHT]
        if bottom_left_index is not None:
            detected_corners[ARUCO_BOTTOM_LEFT_ID] = corners[bottom_left_index][0][ARUCO_CORNER_BOTTOM_RIGHT]
        if bottom_right_index is not None:
            detected_corners[ARUCO_BOTTOM_RIGHT_ID] = corners[bottom_right_index][0][ARUCO_CORNER_BOTTOM_LEFT]
        if top_right_index is not None:
            detected_corners[ARUCO_TOP_RIGHT_ID] = corners[top_right_index][0][ARUCO_CORNER_TOP_LEFT]

        # Estimate one missing fourth position if necessary
        missing_corner_index = index_of_none(detected_corners)
        if missing_corner_index is not None:
            # Use diagonal to calculate missing position
            # Note that this only works if the scanned image is (nearly) a parallelogram (which is usually the case.
            # Alternative approaches:
            # - Use all corners of the three detected aruco markers to find the perspective transformation
            # - Use aruco candidates near the estimated position
            # Two points on the original horizontal border of the grading table
            horizontal_point1 = detected_corners[(missing_corner_index + 1 + (missing_corner_index % 2)) % 4]
            horizontal_point2 = detected_corners[(missing_corner_index + 2 + (missing_corner_index % 2)) % 4]
            # Point on the other horizontal border, with the not detected aruco marker
            third_point = detected_corners[(missing_corner_index + 3 + (missing_corner_index % 2) * 2) % 4]
            third_point_mirrored = mirror_point_across_line(third_point, horizontal_point1, horizontal_point2)
            # Now calculate the missing point by adding the diagonal to the point on the opposite corner
            detected_corners[missing_corner_index] = horizontal_point2 + (horizontal_point1 - third_point_mirrored)

        src_points = np.array(detected_corners, dtype="float32")

        height = max(np.linalg.norm(src_points[0] - src_points[1]), np.linalg.norm(src_points[2] - src_points[3]))
        width = max(np.linalg.norm(src_points[0] - src_points[3]), np.linalg.norm(src_points[1] - src_points[2]))
        destination_points = np.array([
            [0, 0],
            [0, height - 1],
            [width - 1, height - 1],
            [width - 1, 0]
        ], dtype="float32")

        M = cv2.getPerspectiveTransform(src_points, destination_points)

        return cv2.warpPerspective(image, M, (int(width), int(height)))
    else:
        logger.warn("Not enough ArUco markers found - this shouldn't happen here")
        return None


def index_of_none(xs: Iterable) -> Optional[int]:
    for i, x in enumerate(xs):
        try:
            if x is None:
                return i
        except ValueError:
            pass
    return None


def mirror_point_across_line(point: np.array, line_point1: np.array, line_point2: np.array) -> np.array:
    # Extract coordinates
    x, y = point[0], point[1]
    x1, y1 = line_point1
    x2, y2 = line_point2

    # Calculate the slope (m) and intercept (c) of the line
    if x2 - x1 != 0:  # Avoid division by zero for vertical lines
        m = (y2 - y1) / (x2 - x1)
        c = y1 - m * x1
    else:  # Vertical line case
        m = None  # Undefined slope

    if m is not None:
        # Calculate the intersection point
        x_i = (m * (y - c) + x) / (m**2 + 1)
        y_i = m * x_i + c
    else:
        # Special case: Vertical line, reflection across x = x1
        x_i = x1
        y_i = y

    # Calculate the mirrored point
    x_mirrored = 2 * x_i - x
    y_mirrored = 2 * y_i - y

    return np.array([x_mirrored, y_mirrored])


def debug_draw_aruco_markers(corners, ids, image):
    detected_image = aruco.drawDetectedMarkers(image.copy(), corners, ids)
    cv2.imshow("Detected ArUco Markers", detected_image)
    cv2.waitKey(0)
    cv2.destroyAllWindows()


def points_from_video(video_path: str, points_xlsx_path: str, achievable_points: list[int]) -> None:
    if video_path.isnumeric():
        frames = extract_frames_interactively(video_path)
    else:
        frames = extract_frames(video_path)

    exams = detect_points(frames, achievable_points)

    for eg in exams:
        eg.write_training_images(Path("corpus"))

    write_to_xlsx(exams, points_xlsx_path)


def write_to_xlsx(exams: list[GradingTable], points_xlsx_path: str) -> None:
    if len(exams) == 0:
        raise ValueError("No exams to write to xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active

    number_of_fields = len(exams[0].points())
    exercise_headers = [f"{EXERCISE_HEADER_PREFIX}{i}" for i in range(1, number_of_fields)]
    sum_headers = [SUM_RECOGNIZED_HEADER,
                  SUM_WORKSHEET_HEADER,
                  "Σ==Σ?"]
    header_line = exams[0].student_data_columns() + exercise_headers + sum_headers + ["Photo"]
    for i, header in enumerate(header_line, start=1):
        ws.cell(row=1, column=i, value=header)
        if header.startswith(EXERCISE_HEADER_PREFIX):
            ws.column_dimensions[get_column_letter(i)].width = 5

    for eg in exams:
        eg.write_line(ws)

    if os.path.exists(points_xlsx_path):
        shutil.copy(points_xlsx_path, f"{points_xlsx_path}~")

    wb.save(points_xlsx_path)

    logger.info(f"Saved {len(exams)} results to {points_xlsx_path}")


@log_execution_time
def detect_points(cover_pages: Dict[int, np.array], achievable_points: list[int]) -> List[GradingTable]:
    grading_tables = []

    def process_cover_page(student_number: int, image: np.array) -> GradingTable:
        cropped = de_skew_and_crop_image(image)
        cv2.imwrite(f"corpus/gradingtable_{student_number}.png", cropped)
        grading_table = GradingTable(cropped, achievable_points, student_number)
        # print(grading_table)
        return grading_table

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_cover_page, student_number, image) for student_number, image in cover_pages.items()]

        for future in tqdm(concurrent.futures.as_completed(futures), desc="Detecting points"):
            grading_tables.append(future.result())

    return sorted(grading_tables, key=lambda eg: eg.student_number)

def draw_qr_code_bounding_box(image: np.array, points: np.array, color: Tuple[int, int, int]) -> np.array:
    if points is not None:
        image = image.copy()
        points = points[0].astype(int)
        for i in range(len(points)):
            cv2.line(image, tuple(points[i]), tuple(points[(i + 1) % len(points)]), color, 2)
    return image


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(f"Usage: {sys.argv[0]} <video_path> <grades.xlsx> <achievable grades>")
        print(f"  e.g. foo.mkv /tmp/grades.xlsx 9,14,4,10")
        sys.exit(1)
    # TODO accept input Matrikelnummer liste and write empty lines where no data detected
    points_from_video(sys.argv[1], sys.argv[2], [int(p) for p in sys.argv[3].split(",")])
