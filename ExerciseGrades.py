from itertools import product
from pathlib import Path
from typing import List, Dict, Tuple

import cv2
import numpy as np
from openpyxl.worksheet.worksheet import Worksheet
from tensorflow.keras.models import load_model

from WorksheetFunctions import column_index_by_title, column_letter_by_title, write_image_to_cell
from training.create_augmented_base_data_set import DIGIT_IMAGE_SIZE

EMPTY_THRESHOLD = 205

PADDING_CELLS_LEFT = 1 # space (= number of cell widths) between upper right corner of ID 0 and start of grading table cells
PADDING_CELLS_RIGHT = 1

CELL_CROP_PADDING = 0.1


def number_of_cells(achievable_points: List[float]) -> int:
    """Does not count the padding cells"""
    return 3 * len(achievable_points) + number_of_sum_cells(achievable_points)


def number_of_sum_cells(achievable_points) -> int:
    return len(str(int(sum(achievable_points)))) + 1 # + 1 for the decimal point

def debug_display_image(image: np.array):
    cv2.imshow("Image", image)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

class Cell:
    model = load_model('0-10-final.keras')

    def __init__(self, allowed_values: List[int], image: np.array):
        self.allowed_values = allowed_values
        half_height = image.shape[0] // 2
        self.primary_image = Cell.preprocess_image(image[:int(half_height * (1 + CELL_CROP_PADDING)), :])
        self.secondary_image = Cell.preprocess_image(image[int(half_height * (1 - CELL_CROP_PADDING)):, :])

    @staticmethod
    def preprocess_image(image: np.array) -> np.array:
        """converts to grayscale, resizes, pads to target size"""
        # TODO move to training script
        height, width = image.shape[:2]
        if height > width:
            padding = (height - width) // 2
            padded_image = cv2.copyMakeBorder(image, 0, 0, padding, padding, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        else:
            padding = (width - height) // 2
            padded_image = cv2.copyMakeBorder(image, padding, padding, 0, 0, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        resized = cv2.resize(padded_image, (DIGIT_IMAGE_SIZE, DIGIT_IMAGE_SIZE))
        gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY)
        return gray

    @staticmethod
    def is_empty(image: np.array) -> bool:
        # note that the input image should already be contrast-normalized
        center = Cell.image_center(image)
        # debug_display_image(center)
        # print(np.mean(center))
        return np.mean(center) > EMPTY_THRESHOLD

    def secondary_is_empty(self) -> bool:
        return self.is_empty(self.secondary_image)

    @staticmethod
    def image_center(image: np.array) -> np.array:
        # note that the original images are not quadratic; we should shove off the black border
        return image[DIGIT_IMAGE_SIZE // 4:DIGIT_IMAGE_SIZE * 3 // 4, int(DIGIT_IMAGE_SIZE * (3/10)):int(DIGIT_IMAGE_SIZE * (7/10))]

    def detect_number(self, use_secondary: bool) -> Dict[int, float]:
        image = self.secondary_image if use_secondary else self.primary_image

        # Reshape the image to match the model's input shape
        input_image = image.reshape(1, DIGIT_IMAGE_SIZE, DIGIT_IMAGE_SIZE, 1)

        predictions = Cell.model.predict(input_image)[0]
        allowed_predictions = {value: predictions[value] for value in self.allowed_values}

        # debug_display_image(image)

        if Cell.is_empty(image) and 0 in self.allowed_values:
            allowed_predictions[0] = .999
            return allowed_predictions
        return allowed_predictions

class ExerciseGrade:
    def __init__(self, max_value: float, images: List[np.array]):
        self.max_value = max_value

        self.cells = []
        # TODO support für dreistellig; braucht Tests;
        if len(images) != 3:
            raise NotImplementedError("Only 3 cells are supported")
        for idx, image in enumerate(images):
            if idx == 2:
                self.cells.append(Cell([0, 5], image))
            elif idx == 1:
                if max_value < 10:
                    self.cells.append(Cell([n for n in range(int(max_value) % 10 + 1)], image))
                else:
                    self.cells.append(Cell([n for n in range(10)], image))
            elif idx == 0:
                self.cells.append(Cell([n for n in range(int(max_value // 10) % 10 + 1)], image))

    def use_secondary(self) -> bool:
        return any([not cell.secondary_is_empty() for cell in self.cells])

    def detect_number(self) -> List[Tuple[float, float]]:
        """sorted List of detected number and probability pairs; most probable detection first"""
        def build_number(digits: List[int]) -> float:
            number = 0
            for digit in digits:
                number *= 10
                number += digit
            return number / 10

        def get_combinations_with_probabilities(dict_list: List[Dict[int, float]]) -> List[Tuple[List[int], float]]:
            keys_combinations = list(product(*[list(d.keys()) for d in dict_list]))

            combinations_with_probabilities = []

            for combination in keys_combinations:
                if build_number(combination) > self.max_value:
                    continue

                probability = 1
                for idx, key in enumerate(combination):
                    probability *= dict_list[idx][key]
                combinations_with_probabilities.append((list(combination), probability))

            return sorted(combinations_with_probabilities, key=lambda ns_p: ns_p[1], reverse=True)

        detection_results = [cell.detect_number(self.use_secondary()) for cell in self.cells]

        probabilities = get_combinations_with_probabilities(detection_results)
        result = [(build_number(digits), p) for digits, p in probabilities if p > 0.1][:5]
        if len(result) == 0:
            print("TODO: no good match found")
            return [(build_number(probabilities[0][0]), probabilities[0][1])]
        return result


    def get_used_image(self) -> np.array:
        concatenated = np.concatenate([cell.secondary_image if self.use_secondary() else cell.primary_image for cell in self.cells], axis=1)
        column_sums = np.sum(concatenated, axis=0)
        non_white_columns = np.where(column_sums < 255 * concatenated.shape[0])[0]
        if non_white_columns.size > 0:
            return concatenated[:, non_white_columns]
        return concatenated


class ExerciseGrades:
    def __init__(self, image: np.array, achievable_points: List[float], student_number: int):
        self.student_number = student_number
        self._get_cells(image, achievable_points)
        self._grades = self._predict_grades()

    def _get_cells(self, image: np.array, achievable_points: List[float]):
        height, width, _ = image.shape
        total_cells = number_of_cells(achievable_points) + PADDING_CELLS_LEFT + PADDING_CELLS_RIGHT
        cell_width = width / total_cells

        slices = []
        for i in range(PADDING_CELLS_LEFT, total_cells - PADDING_CELLS_RIGHT):
            start_x = int(i * cell_width - CELL_CROP_PADDING * cell_width)
            end_x = int((i + 1) * cell_width + CELL_CROP_PADDING * cell_width)
            slice_img = image[:, start_x:end_x]
            lower_half_slice = slice_img[height // 2:, :]
            # debug_display_image(lower_half_slice)
            slices.append(lower_half_slice)

        self.exercise_grades = []
        for exercise_idx, exercise_points in enumerate(achievable_points):
            self.exercise_grades.append(ExerciseGrade(exercise_points, slices[3 * exercise_idx:3 * exercise_idx + 3]))

        self.sum = ExerciseGrade(sum(achievable_points), slices[-number_of_sum_cells(achievable_points):])


    def _predict_grades(self) -> List[float]:

        def get_combinations_with_probabilities(list_list: List[List[Tuple[float, float]]]) -> List[Tuple[List[float], float]]:
            keys_combinations = list(product(*[list(map(lambda x: x[0], lst)) for lst in list_list]))

            combinations_with_probabilities = []

            for combination in keys_combinations:
                probability = 1
                for idx, key in enumerate(combination):
                    probability *= next(p for k, p in list_list[idx] if k == key)
                combinations_with_probabilities.append((list(combination), probability))

            return sorted(combinations_with_probabilities, key=lambda ns_p: ns_p[1], reverse=True)

        def best_result() -> List[float]:
            detection_results = get_combinations_with_probabilities(
                [exercise_grade.detect_number() for exercise_grade in self.exercise_grades + [self.sum]])
            matching_detection_results = [(grades, p) for grades, p in detection_results if sum(grades[:-1]) == grades[-1]]
            if len(matching_detection_results) > 0:
                return matching_detection_results[0][0]
            print("todo sum not matching:", self.student_number, detection_results)
            return detection_results[0][0]

        prediction = best_result()
        self.predicted_grades = prediction[:-1]
        self.predicted_sum = prediction[-1]
        self.predicted_sum_matches = sum(self.predicted_grades) == self.predicted_sum
        if not self.predicted_sum_matches:
            print("TODO: sum not matching")
        return self.predicted_grades + [self.predicted_sum]

    def grades(self) -> List[float]:
        return self._grades

    def __repr__(self) -> str:
        return str(self.student_number) + "," + ",".join([str(grade) for grade in self.grades()]) + "," + str(self.predicted_sum_matches)

    def write_training_images(self, directory: Path) -> None:
        for idx, exercise_grade in enumerate(self.exercise_grades + [self.sum], start=1):
            for cell_idx, cell in enumerate(exercise_grade.cells):
                image = cell.primary_image
                if exercise_grade.use_secondary():
                    image = cell.secondary_image
                if not Cell.is_empty(image):
                    cv2.imwrite(directory / f"{self.student_number}_{idx}_{cell_idx}.png", image)


    def write_line(self, ws: Worksheet) -> None:
        target_row = ws.max_row + 1

        ws.cell(row=target_row, column=column_index_by_title(ws, "Matrikelnummer"), value=self.student_number)

        exercises_points = self.grades()[:-1]
        for exercise_number, p in enumerate(exercises_points, start=1):
            ws.cell(row=target_row, column=column_index_by_title(ws, f"A{exercise_number}"), value=p)

        sum_points = self.grades()[-1]
        ws.cell(row=target_row, column=column_index_by_title(ws, "Σ (erkannt)"), value=sum_points)

        first_exercise_cell = f"{column_letter_by_title(ws, 'A1')}{target_row}"
        last_exercise_cell = f"{column_letter_by_title(ws, f'A{len(exercises_points)}')}{target_row}"
        sum_formula = f"=SUM({first_exercise_cell}:{last_exercise_cell})"
        ws.cell(row=target_row, column=column_index_by_title(ws, "Σ (von Worksheet berechnet)"), value=sum_formula)

        sum_matches_formula = f"={column_letter_by_title(ws, 'Σ (erkannt)')}{target_row}={column_letter_by_title(ws, 'Σ (von Worksheet berechnet)')}{target_row}"
        ws.cell(row=target_row, column=column_index_by_title(ws, "Σ==Σ?"), value=sum_matches_formula)

        for idx, exercise_grade in enumerate(self.exercise_grades, start=column_index_by_title(ws, 'A1-Bild')):
            write_image_to_cell(ws, exercise_grade.get_used_image(), target_row, idx)
        write_image_to_cell(ws, self.sum.get_used_image(), target_row, column_index_by_title(ws, 'Σ-Bild'))
